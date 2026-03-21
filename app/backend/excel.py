import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import Alignment
import my_gpt

class Excel:

    # I just added this var here
    block = []
    numpy_arr = []
    wb = None
    ws = None
    language = ""
    first_cell_range = ""
    seconds_cell_range = ""
    def __init__(self,file, first_cell_range,second_cell_range): #cell_range example would be A1:B10
        
        input_buf = BytesIO(file) 
        self.wb = load_workbook(input_buf)
        self.ws = self.wb.active
        self.first_cell_range = first_cell_range
        self.seconds_cell_range = second_cell_range
        #Get the column to translate using pandas and convert it into a numpy array
        input_buf.seek(0)  
        df = pd.read_excel(input_buf, index_col=None, na_values=['NA'])
        self.numpy_arr = pd.DataFrame(df)
        print(self.numpy_arr)
       
    def translate(self, language):

        # Parse start/end cells
        start_col_letter = ''.join(c for c in self.first_cell_range if not c.isdigit())
        start_row = int(''.join(c for c in self.first_cell_range if c.isdigit()))

        end_col_letter = ''.join(c for c in self.seconds_cell_range if not c.isdigit())
        end_row = int(''.join(c for c in self.seconds_cell_range if c.isdigit()))

        input_col_letter = start_col_letter     
        output_col_letter = "D"                  

        # Collect text to translate
        self.block = []
        for row_num in range(start_row, end_row + 1):
            cell_value = self.ws[f"{input_col_letter}{row_num}"].value
            if cell_value:
                self.block.append(cell_value)
            else:
                self.block.append("")

        # Call AI
        ai = my_gpt.AssistedIntelligent(self.block, language)
        translated_text = ai.ask()

        # Write translations into column D 
        for i, translation in enumerate(translated_text, start=start_row):
            self.ws[f"{output_col_letter}{i}"] = translation

        # Save file
        output_buf = BytesIO()
        self.wb.save(output_buf)

        with open("translated_output.xlsx", "wb") as f:
            f.write(output_buf.getvalue())
            

