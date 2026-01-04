from openpyxl import Workbook, load_workbook
from io import BytesIO

# Create a test workbook

class ExcelTest:

    def __init__(self,file): 

        input_buf = BytesIO(file)
        wb = load_workbook(input_buf)
        ws = wb.active
        ws["A7"] = "Testing"
        output_buf = BytesIO()
        wb.save(output_buf)

        # ! This is my idea but not sure if it works 
        df = pd.read_excel(input_buf, index_col=None, na_values=['NA'], usecols="A,C:AA")
        print(df)


        with open("out.xlsx", "wb") as f:
            f.write(output_buf.getvalue())
        